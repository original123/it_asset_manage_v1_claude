"""审计日志路由"""
from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from datetime import datetime
import io
from app.models import AuditLog
from app.extensions import db
from app.utils import (
    api_response, admin_required, paginate_query
)

audit_logs_bp = Blueprint('audit_logs', __name__)


@audit_logs_bp.route('', methods=['GET'])
@jwt_required()
@admin_required
def list_audit_logs():
    """获取审计日志列表"""
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 20, type=int)
    user_id = request.args.get('user_id', type=int)
    action = request.args.get('action')
    resource_type = request.args.get('resource_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    keyword = request.args.get('keyword', '').strip()

    query = AuditLog.query

    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if action:
        query = query.filter(AuditLog.action == action)
    if resource_type:
        query = query.filter(AuditLog.resource_type == resource_type)
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(AuditLog.created_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date)
            query = query.filter(AuditLog.created_at <= end_dt)
        except ValueError:
            pass
    if keyword:
        query = query.filter(
            db.or_(
                AuditLog.resource_name.ilike(f'%{keyword}%'),
                AuditLog.username.ilike(f'%{keyword}%'),
            )
        )

    query = query.order_by(AuditLog.created_at.desc())
    result = paginate_query(query, page, page_size)

    return api_response(
        [log.to_dict() for log in result['items']],
        pagination=result['pagination']
    )


@audit_logs_bp.route('/<int:id>', methods=['GET'])
@jwt_required()
@admin_required
def get_audit_log(id):
    """获取审计日志详情"""
    log = AuditLog.query.get_or_404(id)
    return api_response(log.to_dict(include_details=True))


@audit_logs_bp.route('/export', methods=['GET'])
@jwt_required()
@admin_required
def export_audit_logs():
    """导出审计日志"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return api_response(None, '导出功能需要安装openpyxl', code=500)

    # 获取筛选条件
    user_id = request.args.get('user_id', type=int)
    action = request.args.get('action')
    resource_type = request.args.get('resource_type')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = AuditLog.query

    if user_id:
        query = query.filter(AuditLog.user_id == user_id)
    if action:
        query = query.filter(AuditLog.action == action)
    if resource_type:
        query = query.filter(AuditLog.resource_type == resource_type)
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(AuditLog.created_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date)
            query = query.filter(AuditLog.created_at <= end_dt)
        except ValueError:
            pass

    logs = query.order_by(AuditLog.created_at.desc()).limit(10000).all()

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '审计日志'

    # 表头
    headers = ['时间', '用户', 'IP地址', '操作类型', '资源类型', '资源名称', '变更内容']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 数据
    action_map = {'create': '创建', 'update': '更新', 'delete': '删除'}
    type_map = {
        'server': '服务器', 'container': '容器', 'service': '服务',
        'gpu': 'GPU', 'datacenter': '机房', 'user': '用户'
    }

    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '')
        ws.cell(row=row, column=2, value=log.username)
        ws.cell(row=row, column=3, value=log.ip_address or '')
        ws.cell(row=row, column=4, value=action_map.get(log.action, log.action))
        ws.cell(row=row, column=5, value=type_map.get(log.resource_type, log.resource_type))
        ws.cell(row=row, column=6, value=log.resource_name or '')
        ws.cell(row=row, column=7, value=str(log.get_changes()) if log.changes else '')

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
