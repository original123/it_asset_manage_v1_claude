"""导入导出路由"""
import io
from datetime import datetime
from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from app.models import Server, Container, Datacenter, Environment
from app.extensions import db
from app.utils import (
    api_response, error_response, get_current_user, admin_required
)

import_export_bp = Blueprint('import_export', __name__)


@import_export_bp.route('/template', methods=['GET'])
@jwt_required()
def download_template():
    """下载导入模板"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return error_response('需要安装openpyxl', 500, 500)

    wb = openpyxl.Workbook()

    # 服务器模板
    ws_servers = wb.active
    ws_servers.title = '服务器'
    server_headers = [
        '服务器名称*', '机房名称*', '环境*', '内网IP*', '外网IP',
        'CPU核数', '内存(GB)', '磁盘(GB)', '操作系统', 'SSH端口',
        'SSH用户', '负责人', '描述'
    ]
    _write_template_sheet(ws_servers, server_headers, [
        ['web-server-01', '北京机房', '生产环境', '192.168.1.100', '8.210.1.100',
         '16', '64', '500', 'CentOS 7.9', '22', 'root', '张三', 'Web服务器']
    ])

    # 容器模板
    ws_containers = wb.create_sheet('容器')
    container_headers = [
        '容器名称*', '服务器名称*', '镜像', 'CPU限制', '内存限制(MB)',
        '状态', '描述'
    ]
    _write_template_sheet(ws_containers, container_headers, [
        ['nginx-01', 'web-server-01', 'nginx:latest', '2', '2048', 'running', 'Nginx容器']
    ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='import_template.xlsx'
    )


def _write_template_sheet(ws, headers, examples):
    """写入模板工作表"""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15


@import_export_bp.route('/import', methods=['POST'])
@jwt_required()
@admin_required
def import_data():
    """导入数据"""
    try:
        import openpyxl
    except ImportError:
        return error_response('需要安装openpyxl', 500, 500)

    user = get_current_user()

    if 'file' not in request.files:
        return error_response('未上传文件', 422, 422)

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return error_response('仅支持Excel文件', 422, 422)

    overwrite = request.form.get('overwrite', 'false').lower() == 'true'

    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        return error_response(f'文件读取失败: {str(e)}', 422, 422)

    results = {
        'servers': {'created': 0, 'updated': 0, 'errors': []},
        'containers': {'created': 0, 'updated': 0, 'errors': []},
    }

    # 导入服务器
    if '服务器' in wb.sheetnames:
        ws = wb['服务器']
        results['servers'] = _import_servers(ws, overwrite)

    # 导入容器
    if '容器' in wb.sheetnames:
        ws = wb['容器']
        results['containers'] = _import_containers(ws, user, overwrite)

    db.session.commit()

    return api_response(results, '导入完成')


def _import_servers(ws, overwrite):
    """导入服务器数据"""
    result = {'created': 0, 'updated': 0, 'errors': []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:  # 跳过空行
            continue

        try:
            name = str(row[0]).strip()
            datacenter_name = str(row[1]).strip() if row[1] else None
            environment_name = str(row[2]).strip() if row[2] else None
            internal_ip = str(row[3]).strip() if row[3] else None

            if not all([name, datacenter_name, environment_name, internal_ip]):
                result['errors'].append(f'行{row_idx}: 必填字段不完整')
                continue

            # 查找机房
            datacenter = Datacenter.query.filter_by(name=datacenter_name).first()
            if not datacenter:
                datacenter = Datacenter(name=datacenter_name)
                db.session.add(datacenter)
                db.session.flush()

            # 查找环境
            environment = Environment.query.filter_by(name=environment_name).first()
            if not environment:
                result['errors'].append(f'行{row_idx}: 环境 "{environment_name}" 不存在')
                continue

            # 查找或创建服务器
            server = Server.query.filter_by(name=name).first()
            if server:
                if overwrite:
                    server.datacenter_id = datacenter.id
                    server.environment_id = environment.id
                    server.internal_ip = internal_ip
                    server.external_ip = str(row[4]).strip() if row[4] else None
                    server.cpu_cores = int(row[5]) if row[5] else None
                    server.memory_gb = int(row[6]) if row[6] else None
                    server.disk_gb = int(row[7]) if row[7] else None
                    server.os_type = str(row[8]).strip() if row[8] else None
                    server.ssh_port = int(row[9]) if row[9] else 22
                    server.ssh_user = str(row[10]).strip() if row[10] else 'root'
                    server.responsible_person = str(row[11]).strip() if row[11] else None
                    server.description = str(row[12]).strip() if row[12] else None
                    result['updated'] += 1
                else:
                    result['errors'].append(f'行{row_idx}: 服务器 "{name}" 已存在')
            else:
                server = Server(
                    name=name,
                    datacenter_id=datacenter.id,
                    environment_id=environment.id,
                    internal_ip=internal_ip,
                    external_ip=str(row[4]).strip() if row[4] else None,
                    cpu_cores=int(row[5]) if row[5] else None,
                    memory_gb=int(row[6]) if row[6] else None,
                    disk_gb=int(row[7]) if row[7] else None,
                    os_type=str(row[8]).strip() if row[8] else None,
                    ssh_port=int(row[9]) if row[9] else 22,
                    ssh_user=str(row[10]).strip() if row[10] else 'root',
                    responsible_person=str(row[11]).strip() if row[11] else None,
                    description=str(row[12]).strip() if row[12] else None,
                )
                db.session.add(server)
                result['created'] += 1

        except Exception as e:
            result['errors'].append(f'行{row_idx}: {str(e)}')

    return result


def _import_containers(ws, user, overwrite):
    """导入容器数据"""
    result = {'created': 0, 'updated': 0, 'errors': []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row[0]:
            continue

        try:
            name = str(row[0]).strip()
            server_name = str(row[1]).strip() if row[1] else None

            if not all([name, server_name]):
                result['errors'].append(f'行{row_idx}: 必填字段不完整')
                continue

            server = Server.query.filter_by(name=server_name).first()
            if not server:
                result['errors'].append(f'行{row_idx}: 服务器 "{server_name}" 不存在')
                continue

            container = Container.query.filter_by(name=name, server_id=server.id).first()
            if container:
                if overwrite:
                    container.image = str(row[2]).strip() if row[2] else None
                    container.cpu_limit = float(row[3]) if row[3] else None
                    container.memory_limit_mb = int(row[4]) if row[4] else None
                    container.status = str(row[5]).strip() if row[5] else 'running'
                    container.description = str(row[6]).strip() if row[6] else None
                    result['updated'] += 1
                else:
                    result['errors'].append(f'行{row_idx}: 容器 "{name}" 已存在')
            else:
                container = Container(
                    name=name,
                    server_id=server.id,
                    owner_id=user.id,
                    image=str(row[2]).strip() if row[2] else None,
                    cpu_limit=float(row[3]) if row[3] else None,
                    memory_limit_mb=int(row[4]) if row[4] else None,
                    status=str(row[5]).strip() if row[5] else 'running',
                    description=str(row[6]).strip() if row[6] else None,
                )
                db.session.add(container)
                result['created'] += 1

        except Exception as e:
            result['errors'].append(f'行{row_idx}: {str(e)}')

    return result


@import_export_bp.route('/export', methods=['GET'])
@jwt_required()
def export_data():
    """导出数据"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return error_response('需要安装openpyxl', 500, 500)

    export_type = request.args.get('type', 'all')  # all, servers, containers

    wb = openpyxl.Workbook()

    # 导出服务器
    if export_type in ['all', 'servers']:
        ws = wb.active
        ws.title = '服务器'
        _export_servers(ws)

    # 导出容器
    if export_type in ['all', 'containers']:
        if export_type == 'containers':
            ws = wb.active
            ws.title = '容器'
        else:
            ws = wb.create_sheet('容器')
        _export_containers(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'export_{export_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _export_servers(ws):
    """导出服务器数据"""
    from openpyxl.styles import Font, PatternFill, Alignment

    headers = ['ID', '名称', '机房', '环境', '内网IP', '外网IP', 'CPU', '内存(GB)',
               '磁盘(GB)', '系统', 'SSH端口', '负责人', '状态', '创建时间']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    servers = Server.query.order_by(Server.name).all()
    for row, server in enumerate(servers, 2):
        ws.cell(row=row, column=1, value=server.id)
        ws.cell(row=row, column=2, value=server.name)
        ws.cell(row=row, column=3, value=server.datacenter.name if server.datacenter else '')
        ws.cell(row=row, column=4, value=server.environment.name if server.environment else '')
        ws.cell(row=row, column=5, value=server.internal_ip)
        ws.cell(row=row, column=6, value=server.external_ip or '')
        ws.cell(row=row, column=7, value=server.cpu_cores)
        ws.cell(row=row, column=8, value=server.memory_gb)
        ws.cell(row=row, column=9, value=server.disk_gb)
        ws.cell(row=row, column=10, value=server.os_type or '')
        ws.cell(row=row, column=11, value=server.ssh_port)
        ws.cell(row=row, column=12, value=server.responsible_person or '')
        ws.cell(row=row, column=13, value=server.status)
        ws.cell(row=row, column=14, value=server.created_at.strftime('%Y-%m-%d %H:%M') if server.created_at else '')


def _export_containers(ws):
    """导出容器数据"""
    from openpyxl.styles import Font, PatternFill, Alignment

    headers = ['ID', '名称', '服务器', '所有者', '镜像', 'CPU限制', '内存限制(MB)',
               '状态', '端口映射', '创建时间']

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    containers = Container.query.order_by(Container.name).all()
    for row, container in enumerate(containers, 2):
        port_mappings = '; '.join([pm.mapping_chain for pm in container.port_mappings])

        ws.cell(row=row, column=1, value=container.id)
        ws.cell(row=row, column=2, value=container.name)
        ws.cell(row=row, column=3, value=container.server.name if container.server else '')
        ws.cell(row=row, column=4, value=container.owner.display_name if container.owner else '')
        ws.cell(row=row, column=5, value=container.image or '')
        ws.cell(row=row, column=6, value=container.cpu_limit)
        ws.cell(row=row, column=7, value=container.memory_limit_mb)
        ws.cell(row=row, column=8, value=container.status)
        ws.cell(row=row, column=9, value=port_mappings)
        ws.cell(row=row, column=10, value=container.created_at.strftime('%Y-%m-%d %H:%M') if container.created_at else '')


# 需要导入openpyxl
try:
    import openpyxl
except ImportError:
    pass
